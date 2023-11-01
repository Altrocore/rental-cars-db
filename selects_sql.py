import psycopg2
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color

def execute_sql_select_and_save_to_excel(connection, script_path, workbook):
    cursor = connection.cursor()
    with open(script_path, 'r') as file:
        sql_script = file.read()
    cursor.execute(sql_script)
    
    rows = cursor.fetchall()
    column_names = [desc[0] for desc in cursor.description]
    df = pd.DataFrame(rows, columns=column_names)
    
    sheet_name = script_path.replace("/", "_")[:-4] 
    sheet_name = sheet_name[:31]  
    ws = workbook.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    black_fill = PatternFill(start_color="000000",
                             end_color="000000",
                             fill_type="solid")
    white_font = Font(color="FFFFFF")
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = black_fill
            cell.font = white_font

    date_columns = ["Date"] 
    if "Date" in df.columns:
        for cell in ws["Date"]:
            cell.number_format = 'YYYY-MM-DD'

    cursor.close()

connection = psycopg2.connect(
    host="localhost",
    database="your_database_name",
    user="postgres",
    password="your_password"
)

sql_select_scripts = [
    "SelectScripts/CarsAvailableForSale.sql",
    "SelectScripts/CustomerPurchaseHistory.sql",
    "SelectScripts/EmployeeDetails.sql",
    "SelectScripts/MaintenanceRecords.sql",
    "SelectScripts/TotalSalesPerBrand.sql",
    "SelectScripts/UpcomingServiceAppointments.sql",
    "SelectScripts/CarModelsForSpecificBrand.sql"
]

wb = Workbook()

wb.remove(wb.active)

for script_path in sql_select_scripts:
    execute_sql_select_and_save_to_excel(connection, script_path, wb)

wb.save('SQL_Select_Results.xlsx')

connection.close()
